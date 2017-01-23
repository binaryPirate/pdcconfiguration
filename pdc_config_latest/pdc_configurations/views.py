from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.template.context_processors import request
from django.core.files.storage import FileSystemStorage
from collections import defaultdict, namedtuple
from django.template import loader
import shutil
import os
import sys
import openpyxl
from openpyxl import load_workbook
import glob
import cmd
from django.db.transaction import commit
reload(sys)
sys.setdefaultencoding('utf8')
from .models import configuration, post, srcdest
from .forms import Selectenv
from openpyxl import load_workbook
from xml.etree.ElementTree import Element, tostring

import xml.etree.ElementTree as ET

from xml.etree.ElementTree import Element, SubElement, Comment, tostring
import xml.dom.minidom

ozdic=defaultdict(list)
dzdic=defaultdict(list)
chargetuple=namedtuple("char_name"," name orizin destination price glid")
chargelist=list()
chrgdict=defaultdict(list)
dict=defaultdict(list)
def index(request):
    prwrdir = os.path.dirname(__file__)

    xl_location=prwrdir + "\media\\xl"
    print("prestn wordking directory" + xl_location)
    fils=FileSystemStorage()
    fils.delete("media/xl/*")
    #filse=fils.listdir("xl/")
    cleanup()
    #filename=fils[1]
    #f=filename[0]
    #os.remove(prwrdir + "\\media\\xl\\" + "*")

#    f=FileSystemStorage(location="mdedia/xl")
#    print("i am inside ndex and" + str(f))
#    f.delete("BRM_MC050_S001_Config_Mobile Delivery_Tech_Release_v2 02.xlsx")

    return render(request,'pdc_configurations/pdc_configuration.html')


def simple_upload(request):
    prwrdir = os.path.dirname(__file__)

    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        fs = FileSystemStorage(location=prwrdir + "/../" +  "/media/xl",)

        #fs = FileSystemStorage()
        filename = fs.save(myfile.name, myfile)
        uploaded_file_url = fs.url(filename)
        return render(request, 'pdc_configurations/simple_upload.html', {
            'uploaded_file_url': uploaded_file_url
        })
    return render(request, 'pdc_configurations/name.html')


def show_configuration(request):
    prwrdir = os.path.dirname(__file__)

    try :
        os.remove(prwrdir +"\\temp_text_file\\Zone_Model.txt")
    except:
        print("no file present")

    try :
        os.remove(prwrdir +"\\temp_text_file\\balance_element.txt")
    except:
        print("no file present")

    try:
       os.remove(prwrdir +"\\temp_text_file\\impact_category.txt")
    except:
         print("no file present")
    try:
        os.remove(prwrdir +"\data_xml\\Balance_Elements.xml")
    except:
        print("no file ")

    try:
        os.remove(prwrdir + "\data_xml\Impact_Category.xml")
    except:
        print("no xml prest")

    try:
        os.remove(prwrdir + "\data_xml\Zone_Model.xml")
    except:
        print("no xml present")

    #files = os.listdir('C:\Users\chankaya.singh\Desktop\python_tool\django\pdc_config\media')
     # this is added later
    #fils=FileSystemStorage(location="media/xl")
    fils=FileSystemStorage()
    filse=fils.listdir("xl/")
    print(filse)
    filename=filse[1]
    f=filename[0]
    print("present working directory is  " + prwrdir)
#    filse = os.listdir('C:\Users\chankaya.singh\Desktop\\repository\pdc_config_latest\pdc_configurations\media\\xl')
    print(str(filse))
#    print("fie is" + str(filse[0]))

    print(str(f))
    #newest = min(glob.iglob('C:\Users\chankaya.singh\Desktop\python_tool\django\pdc_config\media\\xl\*'), key=os.path.getctime)
    #wb=load_workbook(newest)
    #ws=wb.worksheets[0]
    #wb=load_workbook("media/xl/"+str(f))
    wb=load_workbook("prwdir/../media/xl/"+str(f))
#    wb=load_workbook("media/xl/BRM_MC050_S001_Config_Mobile Delivery_Tech_Release_v2 02.xlsx")
    #wb=load_workbook("C:\Users\chankaya.singh\Desktop\\repository\pdc_config_latest\pdc_configurations\media\\xl")
    sheets=wb.get_sheet_names()
    configuration.objects.all().delete()

    for s in sheets:
        configuration.objects.create(config=s,status="pending")

    template = loader.get_template('pdc_configurations/show_configuration.html')
    config=configuration.objects.all()

    context={
          'config':config,
        }
    return HttpResponse(template.render(context,request))


 ####################################call pdc configuration ##################################

def call_jarvis(request, config_id):
    config_name=get_object_or_404(configuration, pk=config_id)
    #files = os.listdir('C:\\Users\\chankaya.singh\\Desktop\\python_tool\\backup_pdc\\latest\\pdc_config_latest\\media')

    #newest = min(glob.iglob('C:\Users\chankaya.singh\Desktop\python_tool\\backup_pdc\latest\pdc_config_latest\media\*'), key=os.path.getctime)

    #cmd="python C:\Users\chankaya.singh\Desktop\python_tool\\backup_pdc\latest\pdc_config_latest\pdc_configurations\pdc_configuration.py" + " " + "\"" + newest + "\" " + " " + "\"" + str(config_name) + "\"" ""
     ###################this section is addes to creade good code###################################
    fils=FileSystemStorage()
    filse=fils.listdir("xl/")
    print(filse[1])
    filename=filse[1]
    f=filename[0]
    print("config name is " + str(config_name))
    if str(config_name) == 'Zone Model':
            get_data(f,str(config_name))
            create_xml_zone_model();

    elif str(config_name)== 'Balance Elements':
            get_data(f,config_name)
            create_balance_element()

    elif str(config_name) == 'Impact Category':
            get_data(f,config_name)
            create_impact_category();
    elif str(config_name) == "USC_Selector":
            get_data(f,config_name)
            createa_usc_selector_voice()
            create_usc_selector_sms()
            create_usc_selector_mms()
            #createa_usc_selector()
    elif str(config_name)=="Mobile Access Usage Charges":
            get_data(f,config_name)
            create_mbl_acc_usg_chrg()
            create_mbl_acc_usg_chargsms()
            create_mbl_acc_usg_chargmms()
        #    create_mbl_acc_usg_chargsms()
    elif str(config_name)=="Zone Map":
            get_data(f,config_name)
            create_zone_map()
    elif str(config_name)=="Charge Selector":
            get_data(f,config_name)
            create_charges()
            create_charge_selector_xml()
    else:
            print("please pass proper configuration name level1")

    cnfg=configuration.objects.get(id=config_id)
    cnfg.status="done"
    cnfg.save()
    #os.system(cmd)
    cred=post.objects.all()
    for s in cred:
        username=s.usrname
        hostname=s.host_name
        password=s.passwd

    srd=srcdest.objects.all()
    for sd in srd:
        src=sd.source
        dst=sd.destination


    print( username + " " + hostname + " " + password + " " + src + " " + dst)
    config_name=str(config_name)
    config_s=config_name.split(" ")


    #pscpcmd="C:\\\"Program Files\"\\putty\\pscp.exe -pw" + " " + password + " " + src + config_s[0] + "_" + config_s[1] + ".xml"  + " " + username + "@" + hostname + ":" + dst
    #pscpcmd="C:\\\"Program Files\"\\putty\\pscp.exe -pw" + " " + password + " " + "C:\Users\chankaya.singh\Desktop\python_tool\django\pdc_config\data_xml\\" + config_s[0] + "_" + config_s[1] + ".xml"  + " " + username + "@" + hostname + ":/opt/brm/brm/portal/7.5/chan"
    #print(pscpcmd)

    #os.system(pscpcmd)

    template = loader.get_template('pdc_configurations/show_configuration.html')
    config=configuration.objects.all()

    context={
          'config':config,
        }
    return HttpResponse(template.render(context,request))



def envmt(request):

    if request.method == 'POST':
        form=Selectenv(request.POST)

        if form.is_valid():
            host_name=form.cleaned_data['host_name']
            usrname=form.cleaned_data['usrname']
            passwd=form.cleaned_data['passwd']
            src=form.cleaned_data['source']
            dst=form.cleaned_data['destination']
            print(host_name + " " + usrname + " " + passwd )
            post.objects.all().delete()
            srcdest.objects.all().delete()
            post.objects.create(host_name=host_name, usrname=usrname, passwd=passwd)
            srcdest.objects.create(source=src, destination=dst)
            print("inside form valid condition")
            return render(request, 'pdc_configurations/name.html', {'form': form})


    else:
         form=Selectenv()

    print("testing for non post thing")
    return render(request, 'pdc_configurations/server.html', {'form': form})




##########################new functions addesd ###################################
def get_data(src,config_name):
    print("inside get_data functions")
    source_file=src
    print(src)
    pwd = os.path.dirname(__file__)
    print("present workgin directory" + pwd)
    file=pwd + "\..\\" + "\media\\xl\\" +str(source_file)
    #fils=FileSystemStorage()
    print ("xl file is" + file)
    print("config name is " + str(config_name ))
    wb=load_workbook(filename=file)

    sheet_ranges=wb[str(config_name)]
    row=sheet_ranges.max_row
    if str(config_name) == 'Zone Model':
        #file=open(pwd +"temp_text_file\Zone_Model.txt","a")
        file=open(pwd + "\\temp_text_file\\Zone_Model.txt","a")
        for r in range(3,row):
            des=str(sheet_ranges['G' + str(r)].value)
            nam=str(sheet_ranges['H' + str(r)].value)
            file.write( des + ";" + nam + "\n")
        file.close()
    elif str(config_name) == 'Balance Elements':
        print("inside balance elements case")
        file=open(pwd + "\\temp_text_file\\balance_element.txt","a")
        for r in range(3,row):
            name=sheet_ranges['B' + str(r)].value

	    pricelistname="Default"

            code=str(sheet_ranges['C' + str(r)].value)

            numericcode=str(sheet_ranges['D' + str(r)].value)
            symbol=str(sheet_ranges['E' + str(r)].value)

            pr1=str(sheet_ranges['O' + str(r)].value)

            tolmin1=str(sheet_ranges['P' + str(r)].value)

            tolmax1=str(sheet_ranges['Q' + str(r)].value)

            tolper1=str(sheet_ranges['R' + str(r)].value)

            round1=str(sheet_ranges['N' + str(r)].value)

            pr2=str(sheet_ranges['x' + str(r)].value)

            tolmin2=str(sheet_ranges['Y' + str(r)].value)

            tolmax2=str(sheet_ranges['Z' + str(r)].value)

            tolper2=str(sheet_ranges['AA' + str(r)].value)

            round2=str(sheet_ranges['W' + str(r)].value)

            pr3=str(sheet_ranges['AG' + str(r)].value)

            tolmin3=str(sheet_ranges['AH' + str(r)].value)

            tolmax3=str(sheet_ranges['AI' + str(r)].value)

            tolper3=str(sheet_ranges['AJ' + str(r)].value)

            round3=str(sheet_ranges['AF' + str(r)].value)

            pr4=str(sheet_ranges['AP' + str(r)].value)

            tolmin4=str(sheet_ranges['AQ' + str(r)].value)

            tolmax4=str(sheet_ranges['AR' + str(r)].value)

            tolper4=str(sheet_ranges['AS' + str(r)].value)

            round4=str(sheet_ranges['AO' + str(r)].value)
            #file.write(name)
            file.write(name + ";" + pricelistname + ";" + code + ";" + numericcode + ";" + symbol + ";" + pr1 + ";" + tolmin1 + ";" + tolmax1 + ";" + tolper1 + ";" + round1 + ";" + pr2 + ";" + tolmin2 + ";" + tolmax2 + ";" + tolper2 + ";" + round2 + ";" + pr3 + ";" + tolmin3 + ";" + tolmax3 + ";" + tolper3 + ";" + round3 + ";" + pr4 + ";" + tolmin4 + ";" + tolmax4 + ";" + tolper4 + ";" + round4 + ";" + "\n")
	file.close()
    elif str(config_name) == 'Impact Category':
         file=open(pwd + "\\temp_text_file\\impact_category.txt","a")
        # file=open("temp_text_file/impact_category.txt","a")
         for r in range(3,row):
             name=str(sheet_ranges['B' + str(r)].value)
             descr=str(sheet_ranges['C' + str(r)].value)
             s=len(descr)
             if descr[0] == "=":
                ind=descr[1:5]
                if descr[s-4:s-1] == "IDD" :
                    descr=str(sheet_ranges[str(ind)].value)+" " + str(descr[s-4:s-1])
                  #  print(descr + " ")
                else:
                    descr=str(sheet_ranges[str(ind)].value)
                   # print(descr + "\n")



             result=name
             file.write(name + ";" + descr + ";" + result + "\n")
         file.close()
        ##########this section is added for usc selector ############################
    elif str(config_name) == 'USC_Selector':
         print("inside usc selector block")
         i=0
         for ro in sheet_ranges:

             i=i+1
             for c in ro:
                #print("indide row loop")
                if c.value=="Name":
						              #pwd=os.getcwd()
						              #print(pwd)
						              file=open(pwd + "\\" + "temp_text_file" +  "\\" +sheet_ranges["B" + str(i+1)].value +".txt","a")
						              #file=open(pwd + "\\temp_text_file\\impact_category.txt","a")
						              file2=open(pwd + "\\" + "temp_text_file" + "\\" +sheet_ranges["B" + str(i+1)].value +"exception.txt","a")
						              #print(i+1)
						              #print(sheet_ranges["B" + str(i+1)].value)
						              #file.write( str(sheet_ranges["B" + str(i+1)].value + "\n"))
						              for r in range(i+4,row):
									           flag=0
									#if str(sheet["B" + str(r)].value) == "Name":
									           if str(sheet_ranges["D" + str(r)].value) == "None":
											                                        print("here we have name")
											                                        file.close()
											                                        file2.close()
											                                        break

									           else:
										             try:
											                zn_mdl=str(sheet_ranges["D" + str(r)].value)
										             except:
											                 print("its an exception")

										             try:
										                	zn_impct=str(sheet_ranges["E" + str(r)].value)
										             except:
											                  print("its a exception")

										             try:
											                usg_cls=str(sheet_ranges["F" + str(r)].value)
										             except:
											                  print("its a exception")
										             try:
											                impc_cat=str(sheet_ranges["N" + str(r)].value)
										             except:
											                  print("its a exception")


										             ln=len(impc_cat)

										             if impc_cat[0]=="=" and impc_cat[1] != "T":
												            line=impc_cat.split("&")[0]
												            s=len(line)
												            ind=line[1:s]
												            #print("index value is " + ind)
												            if impc_cat[ln-6:ln-1]!="Telia" and impc_cat[ln-6:ln-1] != "Other Network":
													                impc_cat=str(sheet_ranges[ind].value)

												            elif impc_cat[ln-6:ln-1]=="Telia":
													                impc_cat=str(sheet_ranges[ind].value)+ " " + "Telia"
												            elif impc_cat[ln-6:ln-1]=="Other Network":
														         try:
																        impc_cat=str(sheet_ranges[ind].value)+ " " + "Other Network"
														         except:
																        print("error")
												            else:
																        #impc_cat=str(sheet_ranges[ind].value)

																        flag=1
																        print("its a exception")
																        file2.write(zn_mdl + ";" + zn_impct + ";" + usg_cls + ";" + impc_cat + ";" + "\n")
										             elif impc_cat[0]=="=" and impc_cat[1] == "T":
												            line=impc_cat.split("&")[0]
												            s=len(line)
												            ind=line[6:s-1]
												        #    print("index value2 is " + ind)
												            impc_cat=str(sheet_ranges[ind].value)+ " " + "Other Network"




									           if flag!=1:

											            file.write(zn_mdl + ";" + zn_impct + ";" + usg_cls + ";" + impc_cat + ";" + "\n")

									           print(sheet_ranges["E" + str(r)].value)

						#file.close()

    elif str(config_name)=="Mobile Access Usage Charges":
         i=1
         #for rows in sheet_ranges:
        # print("inside maus charges")
         for i in range(3,row):
                i=i+1
             #for c in rows:


                 #if c.value=="Selector":
                if str(sheet_ranges["A" + str(i)].value)=="Selector":
                    #print(" coulumn name is " + str(sheet_ranges["A" + str(i)].value) +" value is " + c.value + " i is " + str(i) )
                    file=open(pwd + "\\" + "temp_text_file" +   "\\" + "Mbl_acc" + str(sheet_ranges["A" + str(i+1)].value) +".txt","a")
                    excpfile=open(pwd + "\\" + "temp_text_file" +   "\\" + "exception_Mbl_acc" + str(sheet_ranges["A" + str(i+1)].value) +".txt","a")
                    for r in range(i+1,row):
                        #print("inside for loop")

                        flag=0
                        if str(sheet_ranges["C" + str(r)].value)=="None":
                            flag=1
                            file.close()
                            break

                        try:
                            impc_cat=str(sheet_ranges["C" + str(r)].value)
                        except:
                            flag=1
                            print(" exception occured")
                        fxd_pr=str(sheet_ranges["D" + str(r)].value)
                        scald_pr=str(sheet_ranges["E" + str(r)].value)
                        glid=str(sheet_ranges["F"+ str(r)].value)


                        if flag==0:
                            file.write(impc_cat + ";" + fxd_pr + ";" + scald_pr + ";" + glid + "\n")

                    if str(sheet_ranges["A" + str(i)].value)=="Charges":
                        print("inside break condition")
                        file.close()
                        excpfile.close()
                        break
                elif str(sheet_ranges["A" + str(i)].value)=="Charges":
                    #file.close()
                    EventName = str(sheet_ranges["A" + str(i+1)].value)
                    print("EventName is" + EventName)
    elif(str(config_name))=="Zone Map":
        # file=open(pwd + "\\temp_text_file\\Zone_map.txt","a")
         for r in range(3,row):
              zkname=str(sheet_ranges['C' + str(r)].value)
              name=str(sheet_ranges['D' + str(r)].value)
              val=str(sheet_ranges['E' + str(r)].value)
              if(str(zkname)=="Origin Zone"):

                  ozdic[name].append(val)
                 # print(ozdic[name])
              elif(str(zkname)=="Destination Zone"):

                  dzdic[name].append(val)
              else:
                  print("we are not comsidering none here")
              result=name
        # for i in ozdic:
        #     for r in ozdic[i]:
        #         print(r)
         #print("destination zone ate ##############")
         #for i in dzdic["Zone1"]:
    #         print(i + "\n")
             # file.write(zkname + ";" + name + ";" + val + "\n")
         #file.close()
    elif(str(config_name))=="Charge Selector":
        for r in range(436,row+1):
            try:
                orgn=str(sheet_ranges['D' + str(r)].value).split(".",1)[1]
                dest=str(sheet_ranges['E' + str(r)].value).split(".",1)[1]
                glid=str(sheet_ranges['I' + str(r)].value)
                pric=str(sheet_ranges['H' + str(r)].value)
                if pric[len(pric)-2:len(pric)] != "kr" :
                    name="Charge"+ " " + str(pric) +"kr"
                else:
                    name="Charge"+ " " + str(pric)

                temp=chargetuple(name,orgn,dest,pric,glid)
        #    print(temp)
            #    print("line number " + str(r))
                chargelist.append(temp)
                chrgdict[name].append(glid)
            except:
                print("found a line wihtout data")

        print("here comea the charge data##################")
        #for i in chargelist:
        #    print i
        #for i in chrgdict:
        #    print("name of charge"+ str(i))
        #    for r in chrgdict[str(i)]:
        #        print( str(r) + "\n")


    else:
         print(" pass the right configuration")


def create_xml_zone_model():


    top=Element('cim:ConfigObjects')
    top.set('xmlns:cim','http://xmlns.oracle.com/communications/platform/model/Config')

    standardZoneModel=SubElement(top,'standardZoneModel')
    standardZoneModel.set('xmlns:mtd','http://xmlns.oracle.com/communications/platform/model/Metadata')
    standardZoneModel.set('xmlns:cim','http://xmlns.oracle.com/communications/platform/model/Config')
    standardZoneModel.set('xmlns:pdc','http://xmlns.oracle.com/communications/platform/model/pricing')
    name=SubElement(standardZoneModel,'name')
    name.text='Standard Zone'
    description=SubElement(standardZoneModel,'description')
    description.text='Standard Zone'
    priceListName=SubElement(standardZoneModel,'priceListName')
    priceListName.text='Default'
    obsolete=SubElement(standardZoneModel,'obsolete')
    obsolete.text='false'
    pwd = os.path.dirname(__file__)

    with open(pwd +"\\temp_text_file\zone_model.txt") as f:
		for line in f:
			desc=line.split(';')[0]
			name=line.split(';')[1]
			name=name.strip()
			desc=desc.strip()
		#	print(name + " " + desc)
			zoneItem=SubElement(standardZoneModel,'zoneItem')
			productName=SubElement(zoneItem,'productName')
			productName.text='*'
			originPrefix=SubElement(zoneItem,'originPrefix')
			originPrefix.text='0'
			destinationPrefix=SubElement(zoneItem,'destinationPrefix')
			destinationPrefix.text=desc
			validFrom=SubElement(zoneItem,'validFrom')
			validFrom.text='0'
			validTo=SubElement(zoneItem,'validTo')
			validTo.text='inf'
			zoneResult=SubElement(zoneItem,'zoneResult')
			zoneName=SubElement(zoneResult,'zoneName')
			zoneName.text=name


    xml_file=open(pwd + "\data_xml\Zone_Model.xml","w")
    #xml_t=xml.dom.minidom.parse(top)
    #pretty_xml=xml_t.toprettyxml()
    #print(pretty_xml)
    xml_file.write(tostring(top))

def create_balance_element():
    top=Element('cim:ConfigObjects')
    top.set('xmlns:cim','http://xmlns.oracle.com/communications/platform/model/Config')
  #  top.set('xmlns:pdc','http://xmlns.oracle.com/communications/platform/model/pricing')
    pwd = os.path.dirname(__file__)
    with open(pwd + "\\temp_text_file\\balance_element.txt") as f:
		for line in f:
                        nam=line.split(';')[0]
                       # nam=nam.strip()
                        cod=line.split(';')[2]
                        numericcode=line.split(';')[3]
                        symbl=line.split(';')[4]
                        pr1=line.split(';')[5]
                        tolmin1=line.split(';')[6]

                        tolmax1=line.split(';')[7]

                        tolper1=line.split(';')[8]
                        if tolmin1 == "0":
                            tolmin1="0.0"

                        if tolper1 == "0":
                            tolper1="0.0"

                        if tolmax1 == "0":
                            tolmax1="0.0"
                        round1=line.split(';')[9]
                        round1=round1.upper()

                        pr2=line.split(';')[10]
                        tolmin2=line.split(';')[11]
                        tolmax2=line.split(';')[12]
                        tolper2=line.split(';')[13]
                        if tolmin2 == "0":
                            tolmin2="0.0"

                        if tolper2 == "0":
                            tolper2="0.0"

                        if tolmax2 == "0":
                            tolmax2="0.0"
                        round2=line.split(';')[14]
                        round2=round2.upper()

                        pr3=line.split(';')[15]
                        tolmin3=line.split(';')[16]
                        tolmax3=line.split(';')[17]
                        tolper3=line.split(';')[18]

                        if tolmin3 == "0":
                            tolmin3="0.0"

                        if tolper3 == "0":
                            tolper3="0.0"

                        if tolmax3 == "0":
                            tolmax3="0.0"
                        round3=line.split(';')[19]
                        round3=round3.upper()

                        pr4=line.split(';')[20]
                        tolmin4=line.split(';')[21]
                        tolmax4=line.split(';')[22]
                        tolper4=line.split(';')[23]
                        if tolmin4 == "0":
                            tolmin4="0.0"

                        if tolper4 == "0":
                            tolper4="0.0"

                        if tolmax4 == "0":
                            tolmax4="0.0"
                        round4=line.split(';')[24]
                        round4=round4.upper()
                        #print(round1.upper() + " " + round2 + " " + round3 + " " + round4 + "\n")

			balanceElements=SubElement(top,'balanceElements')
			balanceElements.set('xmlns:mtd','http://xmlns.oracle.com/communications/platform/model/Metadata')
			balanceElements.set('xmlns:cim','http://xmlns.oracle.com/communications/platform/model/Config')
			balanceElements.set('xmlns:pdc','http://xmlns.oracle.com/communications/platform/model/pricing')
			name=SubElement(balanceElements,'name')
			name.text=nam;
			description=SubElement(balanceElements,'description')
			description.text=nam;
			priceListName=SubElement(balanceElements,'priceListName')
			priceListName.text="default"
			obsolete=SubElement(balanceElements,'obsolete')
			obsolete.text="false"
			code=SubElement(balanceElements,'code')
			code.text=cod
			numericCode=SubElement(balanceElements,'numericCode')
			numericCode.text=numericcode
			symbol=SubElement(balanceElements,'symbol')
			symbol.text=symbl
			transientElement=SubElement(balanceElements,'transientElement')
			transientElement.text="false"
			foldable=SubElement(balanceElements,'foldable')
			foldable.text="true"
			counter=SubElement(balanceElements,'counter')
			counter.text="false"
			roundingRules=SubElement(balanceElements,'roundingRules')
			precision=SubElement(roundingRules,'precision')
			precision.text=pr1
			toleranceMin=SubElement(roundingRules,'toleranceMin')
			toleranceMin.text=tolmin1
			toleranceMax=SubElement(roundingRules,'toleranceMax')
			toleranceMax.text=tolmax1
			tolerancePercentage=SubElement(roundingRules,'tolerancePercentage')
			tolerancePercentage.text=tolper1
			roundingMode=SubElement(roundingRules,'roundingMode')
			roundingMode.text=round1
			processingStage=SubElement(roundingRules,'processingStage')
                        processingStage.text='CHARGING'

                        roundingRules=SubElement(balanceElements,'roundingRules')
			precision=SubElement(roundingRules,'precision')
			precision.text=pr2
			toleranceMin=SubElement(roundingRules,'toleranceMin')
			toleranceMin.text=tolmin2
			toleranceMax=SubElement(roundingRules,'toleranceMax')
			toleranceMax.text=tolmax2
			tolerancePercentage=SubElement(roundingRules,'tolerancePercentage')
			tolerancePercentage.text=tolper2
			roundingMode=SubElement(roundingRules,'roundingMode')
			roundingMode.text=round2
			processingStage=SubElement(roundingRules,'processingStage')
                        processingStage.text='ALTERATION'

			roundingRules=SubElement(balanceElements,'roundingRules')
			precision=SubElement(roundingRules,'precision')
			precision.text=pr3
			toleranceMin=SubElement(roundingRules,'toleranceMin')
			toleranceMin.text=tolmin3
			toleranceMax=SubElement(roundingRules,'toleranceMax')
			toleranceMax.text=tolmax3
			tolerancePercentage=SubElement(roundingRules,'tolerancePercentage')
			tolerancePercentage.text=tolper3
			roundingMode=SubElement(roundingRules,'roundingMode')
			roundingMode.text=round3
			processingStage=SubElement(roundingRules,'processingStage')
                        processingStage.text='TAXATION'

			roundingRules=SubElement(balanceElements,'roundingRules')
			precision=SubElement(roundingRules,'precision')
			precision.text=pr4
			toleranceMin=SubElement(roundingRules,'toleranceMin')
			toleranceMin.text=tolmin4
			toleranceMax=SubElement(roundingRules,'toleranceMax')
			toleranceMax.text=tolmax4
			tolerancePercentage=SubElement(roundingRules,'tolerancePercentage')
			tolerancePercentage.text=tolper4
			roundingMode=SubElement(roundingRules,'roundingMode')
			roundingMode.text=round4
			processingStage=SubElement(roundingRules,'processingStage')
                        processingStage.text='AR'

			consumptionRule=SubElement(balanceElements,'consumptionRule')
			consumptionRule.text="NONE"


	#xml=xml.dom.minidom.parse(top)
	#pretty_xml_as_string=xml.toprettyxnl()
	#print(pretty_xml_as_string)

    xml_file=open(pwd + "\data_xml\Balance_Elements.xml","w")
    xml_file.write(tostring(top))

# function to create xml for impact category
def create_impact_category():
    top=Element('cim:ConfigObjects')
    top.set('xmlns:cim','http://xmlns.oracle.com/communications/platform/model/Config')
    pwd = os.path.dirname(__file__)
    with open(pwd + "\\temp_text_file\impact_category.txt") as f:
        for line in f:

            nam=line.split(";")[0]
            descr=line.split(";")[1]
            reslt=line.split(";")[2]
            zoneResultConfiguration=SubElement(top,'zoneResultConfiguration')
            zoneResultConfiguration.set('xmlns:mtd','http://xmlns.oracle.com/communications/platform/model/Metadata')
            zoneResultConfiguration.set('xmlns:cim','http://xmlns.oracle.com/communications/platform/model/Config')
            zoneResultConfiguration.set('xmlns:pdc','http://xmlns.oracle.com/communications/platform/model/pricing')
            name=SubElement(zoneResultConfiguration,'name')
            name.text=nam
            description=SubElement(zoneResultConfiguration,'description')
            description.text=descr
            priceListName=SubElement(zoneResultConfiguration,'priceListName')
            priceListName.text="Default"
            obsolete=SubElement(zoneResultConfiguration,'obsolete')
            obsolete.text="false"
            result=SubElement(zoneResultConfiguration,'result')
            result.text=reslt
            resultType=SubElement(zoneResultConfiguration,'resultType')
            resultType.text="BASE_RESULT"
        xml_file=open(pwd + "\data_xml\Impact_Category.xml","w")

        xml_file.write(tostring(top))

def createa_usc_selector_voice():
    top=Element('pdc:PricingObjectsJXB')
    top.set('xmlns:pdc','http://xmlns.oracle.com/communications/platform/model/pricing')
    pwd = os.path.dirname(__file__)
    uscSelector=SubElement(top,'uscSelector')
    uscSelector.set('xmlns:pdc','http://xmlns.oracle.com/communications/platform/model/pricing')
    name=SubElement(uscSelector,'name')
    name.text="tesitng name"
    pricingProfileName=SubElement(uscSelector,'pricingProfileName')
    pricingProfileName.text="Convergent Usage"
    priceListName=SubElement(uscSelector,'priceListName')
    priceListName.text="Default"
    stereoType=SubElement(uscSelector,'stereoType')
    stereoType.text="USC_MAP"
    productSpecName=SubElement(uscSelector,'productSpecName')
    productSpecName.text="TelcoGsm"
    eventSpecName=SubElement(uscSelector,'eventSpecName')
    eventSpecName.text='EventDelayedSessionTelcoGsm'
    validityPeriod=SubElement(uscSelector,'validityPeriod')
    validFrom=SubElement(validityPeriod,'validFrom')
    validFrom.text="0"
    with open(pwd + '\\temp_text_file\\Mobile Access Local Voice.txt') as f:
        for line in f:
            impact_cat=line.split(";")[3]
            zone_imp_cat=line.split(";")[1]
            usg_clas=line.split(";")[2]
            if usg_clas == "None":
                usg_clas=".*"

            rule=SubElement(validityPeriod,'rule')
            ruleData=SubElement(rule,'ruleData')
            key=SubElement(ruleData,'key')
            key.text="ZONE_MODEL"
            value=SubElement(ruleData,'value')
            value.text="Standard Zone"
            result=SubElement(rule,'result')
            resultName=SubElement(result,'resultName')
            resultName.text=impact_cat
            fieldToValueExpression=SubElement(rule,'fieldToValueExpression')
            operation=SubElement(fieldToValueExpression,'operation')
            operation.text="REGEX"
            seperator=SubElement(fieldToValueExpression,'seperator')
            seperator.text=";"
            fieldName=SubElement(fieldToValueExpression,'fieldName')
            fieldName.text="EventDelayedSessionTelcoGsm.RETAIL_ZONE"
            fieldValue=SubElement(fieldToValueExpression,'fieldValue')
            fieldValue.text=zone_imp_cat
            fieldKind=SubElement(fieldToValueExpression,'fieldKind')
            fieldKind.text="EVENT_SPEC_FIELD"
            #############
            fieldToValueExpression=SubElement(rule,'fieldToValueExpression')
            operation=SubElement(fieldToValueExpression,'operation')
            operation.text="REGEX"
            seperator=SubElement(fieldToValueExpression,'seperator')
            seperator.text=";"
            fieldName=SubElement(fieldToValueExpression,'fieldName')
            fieldName.text="EventDelayedSessionTelcoGsm.TELCO_INFO.USAGE_CLASS"
            fieldValue=SubElement(fieldToValueExpression,'fieldValue')
            fieldValue.text=usg_clas
            fieldKind=SubElement(fieldToValueExpression,'fieldKind')
            fieldKind.text="EVENT_SPEC_FIELD"
            ##################
            fieldToValueExpression=SubElement(rule,'fieldToValueExpression')
            operation=SubElement(fieldToValueExpression,'operation')
            operation.text="REGEX"
            seperator=SubElement(fieldToValueExpression,'seperator')
            seperator.text=";"
            fieldName=SubElement(fieldToValueExpression,'fieldName')
            fieldName.text="EventDelayedSessionTelcoGsm.SERVICE_CLASS"
            fieldValue=SubElement(fieldToValueExpression,'fieldValue')
            fieldValue.text=".*"
            fieldKind=SubElement(fieldToValueExpression,'fieldKind')
            fieldKind.text="EVENT_SPEC_FIELD"
            ###########################
            fieldToValueExpression=SubElement(rule,'fieldToValueExpression')
            operation=SubElement(fieldToValueExpression,'operation')
            operation.text="REGEX"
            seperator=SubElement(fieldToValueExpression,'seperator')
            seperator.text=";"
            fieldName=SubElement(fieldToValueExpression,'fieldName')
            fieldName.text="EventDelayedSessionTelcoGsm.SERVICE_CODE"
            fieldValue=SubElement(fieldToValueExpression,'fieldValue')
            fieldValue.text=".*"
            fieldKind=SubElement(fieldToValueExpression,'fieldKind')
            fieldKind.text="EVENT_SPEC_FIELD"

        xml_file=open(pwd + "\data_xml\USC_VOICE.xml","w")

        xml_file.write(tostring(top))




def create_mbl_acc_usg_chrg():
    pwd = os.path.dirname(__file__)
    top=Element('chargeRatePlan')
    top.set('xmlns:pdc','http://xmlns.oracle.com/communications/platform/model/pricing')
    name=SubElement(top,'name')
    name.text="here put the name"
    pricingProfileName=SubElement(top,'pricingProfileName')
    pricingProfileName.text="Convergent Usage"
    priceListName=SubElement(top,'priceListName')
    priceListName.text="Default"
    applicableRums=SubElement(top,'applicableRums')
    applicableRums.text="Duration"
    applicableQuantity=SubElement(top,'applicableQuantity')
    applicableQuantity.text="ORIGINAL"
    taxTime=SubElement(top,'taxTime')
    taxTime.text="NONE"
    todMode=SubElement(top,'todMode')
    todMode.text="START_TIME"
    applicableQtyTreatment=SubElement(top,'applicableQtyTreatment')
    applicableQtyTreatment.text="CONTINUOUS"
    permittedName=SubElement(top,'permittedName')
    permittedName.text="TelcoGsm"
    permittedType=SubElement(top,'permittedType')
    permittedType.text="PRODUCT"
    eventName=SubElement(top,'eventName')
    eventName.text="EventDelayedSessionTelcoGsm"
    cycleFeeFlag=SubElement(top,'cycleFeeFlag')
    cycleFeeFlag.text="0"
    billOffset=SubElement(top,'billOffset')
    billOffset.text="0"
    subscriberCurrency=SubElement(top,'subscriberCurrency')
    currencyCode=SubElement(subscriberCurrency,'currencyCode')
    currencyCode.text="SEK"
    applicableRum=SubElement(subscriberCurrency,'applicableRum')
    applicableRumName=SubElement(applicableRum,'applicableRumName')
    applicableRumName.text="Duration"
    minQuantity=SubElement(applicableRum,'minQuantity')
    minQuantity.text="1.0"
    minQuantityUnit=SubElement(applicableRum,'minQuantityUnit')
    minQuantityUnit.text="NONE"
    incrementQuantity=SubElement(applicableRum,'incrementQuantity')
    incrementQuantity.text="0.0"
    incrementQuantityUnit=SubElement(applicableRum,'incrementQuantityUnit')
    incrementQuantityUnit.text="NONE"
    roundingMode=SubElement(applicableRum,'roundingMode')
    roundingMode.text="NEAREST"
    crpRelDateRange=SubElement(applicableRum,'crpRelDateRange')
    absoluteDateRange=SubElement(crpRelDateRange,'absoluteDateRange')
    startDate=SubElement(absoluteDateRange,'startDate')
    startDate.text="0"
    endDate=SubElement(absoluteDateRange,'endDate')
    endDate.text="inf"
    enhancedZoneModel=SubElement(crpRelDateRange,'enhancedZoneModel')
    zoneModelName=SubElement(enhancedZoneModel,'zoneModelName')
    zoneModelName.text="Standard Zone"
    uscModelName=SubElement(enhancedZoneModel,'uscModelName')
    uscModelName.text="Test SM USC Selector"
    with open(pwd + '\\temp_text_file\\Mbl_accMobile Access Local Voice.txt') as f:
        for line in f:
            nme=str(line.split(";")[0])

            fp=line.split(";")[1]
            sp=line.split(";")[2]
            gld=line.split(";")[3]
        #    print( str(nme) + "    " + str(sp)  +" " + str(fp) + " " + str(gld) + "\n")
            results=SubElement(enhancedZoneModel,"results")
            name=SubElement(results,"name")
            name.text=str(nme)
            crpCompositePopModel=SubElement(results,"crpCompositePopModel")
            name_crp=SubElement(crpCompositePopModel,"name")
            name_crp.text="Pricing"
            usageChargePopModel=SubElement(crpCompositePopModel,"usageChargePopModel")
            priceTier=SubElement(usageChargePopModel,"priceTier")
            distributionMethod=SubElement(priceTier,"distributionMethod")
            distributionMethod.text="FROM_BAL_IMPACT"
            tierBasis=SubElement(priceTier,"tierBasis")
            rumTierExpression=SubElement(tierBasis,"rumTierExpression")
            enforceCreditLimit=SubElement(priceTier,"enforceCreditLimit")
            enforceCreditLimit.text="false"
            rumName=SubElement(priceTier,"rumName")
            rumName.text="Duration"
            priceTierValidityPeriod=SubElement(priceTier,"priceTierValidityPeriod")

            lowerBound=SubElement(priceTierValidityPeriod,"lowerBound")
            lowerBound.text="0"
            validFrom=SubElement(priceTierValidityPeriod,"validFrom")
            validFrom.text="0"
            priceTierRange=SubElement(priceTierValidityPeriod,"priceTierRange")
            upperBound=SubElement(priceTierRange,"upperBound")
            upperBound.text="NO_MAX"

            if fp=="-" or fp == "None":
                print("")

            #elif fp !="-" or fp!="None":
            else:
                fixedCharge=SubElement(priceTierRange,"fixedCharge")
                price=SubElement(fixedCharge,"price")
                price.text=str(fp)
                unitOfMeasure=SubElement(fixedCharge,"unitOfMeasure")
                unitOfMeasure.text="NONE"

                balanceElementNumCode=SubElement(fixedCharge,"balanceElementNumCode")
                balanceElementNumCode.text="752"
                discountable=SubElement(fixedCharge,"discountable")
                discountable.text="true"
                priceType=SubElement(fixedCharge,"priceType")
                priceType.text="CONSUMPTION"
                glid=SubElement(fixedCharge,"glid")
                glid.text=gld

            if sp=="-" or sp == "None":
                print("")

            else:
            #if sp!="-" or sp != "None":
                scaledCharge=SubElement(priceTierRange,"scaledCharge")
                price=SubElement(scaledCharge,"price")
                price.text=str(sp)
                unitOfMeasure=SubElement(scaledCharge,"unitOfMeasure")
                unitOfMeasure.text="MINUTES"

                balanceElementNumCode=SubElement(scaledCharge,"balanceElementNumCode")
                balanceElementNumCode.text="752"
                discountable=SubElement(scaledCharge,"discountable")
                discountable.text="true"
                priceType=SubElement(scaledCharge,"priceType")
                priceType.text="CONSUMPTION"
                glid=SubElement(scaledCharge,"glid")
                glid.text=gld
                incrementStep=SubElement(scaledCharge,"incrementStep")
                incrementStep.text="1"
                incrementRounding=SubElement(scaledCharge,"incrementRounding")
                incrementRounding.text="NONE"



            applicableQuantity=SubElement(priceTier,"applicableQuantity")
            applicableQuantity.text="ORIGINAL"
    xml_file=open(pwd + "\data_xml\mobile_usage_voice.xml","w")
    xml_file.write(tostring(top))



def create_usc_selector_sms():
    top=Element('pdc:PricingObjectsJXB')
    top.set('xmlns:pdc','http://xmlns.oracle.com/communications/platform/model/pricing')
    pwd = os.path.dirname(__file__)
    uscSelector=SubElement(top,'uscSelector')
    uscSelector.set('xmlns:pdc','http://xmlns.oracle.com/communications/platform/model/pricing')
    name=SubElement(uscSelector,'name')
    name.text="tesitng name"
    pricingProfileName=SubElement(uscSelector,'pricingProfileName')
    pricingProfileName.text="Convergent Usage"
    priceListName=SubElement(uscSelector,'priceListName')
    priceListName.text="Default"
    stereoType=SubElement(uscSelector,'stereoType')
    stereoType.text="USC_MAP"
    productSpecName=SubElement(uscSelector,'productSpecName')
    productSpecName.text="TelcoGsm"
    eventSpecName=SubElement(uscSelector,'eventSpecName')
    eventSpecName.text='EventDelayedSessionTelcoGsmSms'
    validityPeriod=SubElement(uscSelector,'validityPeriod')
    validFrom=SubElement(validityPeriod,'validFrom')
    validFrom.text="0"
    with open(pwd + '\\temp_text_file\\Mobile Access Local SMS.txt') as f:
        for line in f:
            impact_cat=str(line.split(";")[3])
            zone_imp_cat=line.split(";")[1]
            usg_clas=line.split(";")[2]
            if usg_clas == "None" or usg_clas== "-":
                usg_clas=".*"

            rule=SubElement(validityPeriod,'rule')
            ruleData=SubElement(rule,'ruleData')
            key=SubElement(ruleData,'key')
            key.text="ZONE_MODEL"
            value=SubElement(ruleData,'value')
            value.text="Standard Zone"
            result=SubElement(rule,'result')
            resultName=SubElement(result,'resultName')
            resultName.text=impact_cat
            fieldToValueExpression=SubElement(rule,'fieldToValueExpression')
            operation=SubElement(fieldToValueExpression,'operation')
            operation.text="REGEX"
            seperator=SubElement(fieldToValueExpression,'seperator')
            seperator.text=";"
            fieldName=SubElement(fieldToValueExpression,'fieldName')
            fieldName.text="EventDelayedSessionTelcoGsmSms.RETAIL_ZONE"
            fieldValue=SubElement(fieldToValueExpression,'fieldValue')
            fieldValue.text=zone_imp_cat
            fieldKind=SubElement(fieldToValueExpression,'fieldKind')
            fieldKind.text="EVENT_SPEC_FIELD"
            #############
            fieldToValueExpression=SubElement(rule,'fieldToValueExpression')
            operation=SubElement(fieldToValueExpression,'operation')
            operation.text="REGEX"
            seperator=SubElement(fieldToValueExpression,'seperator')
            seperator.text=";"
            fieldName=SubElement(fieldToValueExpression,'fieldName')
            fieldName.text="EventDelayedSessionTelcoGsmSms.USAGE_CLASS"
            fieldValue=SubElement(fieldToValueExpression,'fieldValue')
            fieldValue.text=usg_clas
            fieldKind=SubElement(fieldToValueExpression,'fieldKind')
            fieldKind.text="EVENT_SPEC_FIELD"
            ##################
            fieldToValueExpression=SubElement(rule,'fieldToValueExpression')
            operation=SubElement(fieldToValueExpression,'operation')
            operation.text="REGEX"
            seperator=SubElement(fieldToValueExpression,'seperator')
            seperator.text=";"
            fieldName=SubElement(fieldToValueExpression,'fieldName')
            fieldName.text="EventDelayedSessionTelcoGsmSms.SERVICE_CLASS"
            fieldValue=SubElement(fieldToValueExpression,'fieldValue')
            fieldValue.text=".*"
            fieldKind=SubElement(fieldToValueExpression,'fieldKind')
            fieldKind.text="EVENT_SPEC_FIELD"
            ###########################
            fieldToValueExpression=SubElement(rule,'fieldToValueExpression')
            operation=SubElement(fieldToValueExpression,'operation')
            operation.text="REGEX"
            seperator=SubElement(fieldToValueExpression,'seperator')
            seperator.text=";"
            fieldName=SubElement(fieldToValueExpression,'fieldName')
            fieldName.text="EventDelayedSessionTelcoGsmSms.SERVICE_CODE"
            fieldValue=SubElement(fieldToValueExpression,'fieldValue')
            fieldValue.text=".*"
            fieldKind=SubElement(fieldToValueExpression,'fieldKind')
            fieldKind.text="EVENT_SPEC_FIELD"

        xml_file=open(pwd + "\data_xml\USC_SMS.xml","w")

        xml_file.write(tostring(top))


    #return HttpResponse("hi this is usc selector")
def create_mbl_acc_usg_chargsms():
    pwd = os.path.dirname(__file__)
    top=Element('chargeRatePlan')
    top.set('xmlns:pdc','http://xmlns.oracle.com/communications/platform/model/pricing')
    name=SubElement(top,'name')
    name.text="here put the name"
    pricingProfileName=SubElement(top,'pricingProfileName')
    pricingProfileName.text="Convergent Usage"
    priceListName=SubElement(top,'priceListName')
    priceListName.text="Default"
    applicableRums=SubElement(top,'applicableRums')
    applicableRums.text="Occurrence"
    applicableQuantity=SubElement(top,'applicableQuantity')
    applicableQuantity.text="ORIGINAL"
    taxTime=SubElement(top,'taxTime')
    taxTime.text="NONE"
    todMode=SubElement(top,'todMode')
    todMode.text="START_TIME"
    applicableQtyTreatment=SubElement(top,'applicableQtyTreatment')
    applicableQtyTreatment.text="CONTINUOUS"
    permittedName=SubElement(top,'permittedName')
    permittedName.text="TelcoGsm"
    permittedType=SubElement(top,'permittedType')
    permittedType.text="PRODUCT"
    eventName=SubElement(top,'eventName')
    eventName.text="EventDelayedSessionTelcoGsmSms"
    cycleFeeFlag=SubElement(top,'cycleFeeFlag')
    cycleFeeFlag.text="0"
    billOffset=SubElement(top,'billOffset')
    billOffset.text="0"
    subscriberCurrency=SubElement(top,'subscriberCurrency')
    currencyCode=SubElement(subscriberCurrency,'currencyCode')
    currencyCode.text="SEK"
    applicableRum=SubElement(subscriberCurrency,'applicableRum')
    applicableRumName=SubElement(applicableRum,'applicableRumName')
    applicableRumName.text="Occurrence"
    minQuantity=SubElement(applicableRum,'minQuantity')
    minQuantity.text="1.0"
    minQuantityUnit=SubElement(applicableRum,'minQuantityUnit')
    minQuantityUnit.text="NONE"
    incrementQuantity=SubElement(applicableRum,'incrementQuantity')
    incrementQuantity.text="0.0"
    incrementQuantityUnit=SubElement(applicableRum,'incrementQuantityUnit')
    incrementQuantityUnit.text="NONE"
    roundingMode=SubElement(applicableRum,'roundingMode')
    roundingMode.text="NEAREST"
    crpRelDateRange=SubElement(applicableRum,'crpRelDateRange')
    absoluteDateRange=SubElement(crpRelDateRange,'absoluteDateRange')
    startDate=SubElement(absoluteDateRange,'startDate')
    startDate.text="0"
    endDate=SubElement(absoluteDateRange,'endDate')
    endDate.text="inf"
    enhancedZoneModel=SubElement(crpRelDateRange,'enhancedZoneModel')
    zoneModelName=SubElement(enhancedZoneModel,'zoneModelName')
    zoneModelName.text="Standard Zone"
    uscModelName=SubElement(enhancedZoneModel,'uscModelName')
    uscModelName.text="Test SMS USC Selector"
    with open(pwd + '\\temp_text_file\\Mbl_accMobile Access Local SMS.txt') as f:
        for line in f:
            nme=str(line.split(";")[0])

            fp=line.split(";")[1]
            sp=line.split(";")[2]
            gld=line.split(";")[3]
        #    print( str(nme) + "    " + str(sp)  +" " + str(fp) + " " + str(gld) + "\n")
            results=SubElement(enhancedZoneModel,"results")
            name=SubElement(results,"name")
            name.text=str(nme)
            crpCompositePopModel=SubElement(results,"crpCompositePopModel")
            name_crp=SubElement(crpCompositePopModel,"name")
            name_crp.text="Pricing"
            usageChargePopModel=SubElement(crpCompositePopModel,"usageChargePopModel")
            priceTier=SubElement(usageChargePopModel,"priceTier")
            distributionMethod=SubElement(priceTier,"distributionMethod")
            distributionMethod.text="FROM_BAL_IMPACT"
            tierBasis=SubElement(priceTier,"tierBasis")
            rumTierExpression=SubElement(tierBasis,"rumTierExpression")
            enforceCreditLimit=SubElement(priceTier,"enforceCreditLimit")
            enforceCreditLimit.text="false"
            rumName=SubElement(priceTier,"rumName")
            rumName.text="Occurrence"
            priceTierValidityPeriod=SubElement(priceTier,"priceTierValidityPeriod")

            lowerBound=SubElement(priceTierValidityPeriod,"lowerBound")
            lowerBound.text="0"
            validFrom=SubElement(priceTierValidityPeriod,"validFrom")
            validFrom.text="0"
            priceTierRange=SubElement(priceTierValidityPeriod,"priceTierRange")
            upperBound=SubElement(priceTierRange,"upperBound")
            upperBound.text="NO_MAX"

            if fp=="-" or fp == "None":
                print("price is not valid ")

            #elif fp !="-" or fp!="None":
            else:
                fixedCharge=SubElement(priceTierRange,"fixedCharge")
                price=SubElement(fixedCharge,"price")
                price.text=str(fp)
                unitOfMeasure=SubElement(fixedCharge,"unitOfMeasure")
                unitOfMeasure.text="NONE"

                balanceElementNumCode=SubElement(fixedCharge,"balanceElementNumCode")
                balanceElementNumCode.text="752"
                discountable=SubElement(fixedCharge,"discountable")
                discountable.text="true"
                priceType=SubElement(fixedCharge,"priceType")
                priceType.text="CONSUMPTION"

                if gld is "None":
                    print("glid is none" + gld + "\n")
                else:
                    glid=SubElement(fixedCharge,"glid")
                    glid.text=gld

            if sp=="-" or sp == "None":
                print("wrong price ")

            else:
            #if sp!="-" or sp != "None":
                scaledCharge=SubElement(priceTierRange,"scaledCharge")
                price=SubElement(scaledCharge,"price")
                price.text=str(sp)
                unitOfMeasure=SubElement(scaledCharge,"unitOfMeasure")
                unitOfMeasure.text="MINUTES"

                balanceElementNumCode=SubElement(scaledCharge,"balanceElementNumCode")
                balanceElementNumCode.text="752"
                discountable=SubElement(scaledCharge,"discountable")
                discountable.text="true"
                priceType=SubElement(scaledCharge,"priceType")
                priceType.text="CONSUMPTION"

                if gld is "None":
                    print("glid is null" + gld +  "\n")
                else:
                    print(" gld is " + gld + "\n")
                    glid=SubElement(scaledCharge,"glid")
                    glid.text=gld

                incrementStep=SubElement(scaledCharge,"incrementStep")
                incrementStep.text="1"
                incrementRounding=SubElement(scaledCharge,"incrementRounding")
                incrementRounding.text="NONE"



            applicableQuantity=SubElement(priceTier,"applicableQuantity")
            applicableQuantity.text="ORIGINAL"
    xml_file=open(pwd + "\data_xml\mobile_usage_SMS.xml","w")
    xml_file.write(tostring(top))


def create_usc_selector_mms():
    top=Element('pdc:PricingObjectsJXB')
    top.set('xmlns:pdc','http://xmlns.oracle.com/communications/platform/model/pricing')
    pwd = os.path.dirname(__file__)
    uscSelector=SubElement(top,'uscSelector')
    uscSelector.set('xmlns:pdc','http://xmlns.oracle.com/communications/platform/model/pricing')
    name=SubElement(uscSelector,'name')
    name.text="tesitng name"
    pricingProfileName=SubElement(uscSelector,'pricingProfileName')
    pricingProfileName.text="Convergent Usage"
    priceListName=SubElement(uscSelector,'priceListName')
    priceListName.text="Default"
    stereoType=SubElement(uscSelector,'stereoType')
    stereoType.text="USC_MAP"
    productSpecName=SubElement(uscSelector,'productSpecName')
    productSpecName.text="TelcoGsm"
    eventSpecName=SubElement(uscSelector,'eventSpecName')
    eventSpecName.text='EventDelayedSessionTelcoGsmSms'
    validityPeriod=SubElement(uscSelector,'validityPeriod')
    validFrom=SubElement(validityPeriod,'validFrom')
    validFrom.text="0"
    with open(pwd + '\\temp_text_file\\Mobile Access Local MMS.txt') as f:
        for line in f:
            impact_cat=line.split(";")[3]
            zone_imp_cat=line.split(";")[1]
            usg_clas=line.split(";")[2]
            if usg_clas == "None" or usg_clas== "-":
                usg_clas=".*"

            rule=SubElement(validityPeriod,'rule')
            ruleData=SubElement(rule,'ruleData')
            key=SubElement(ruleData,'key')
            key.text="ZONE_MODEL"
            value=SubElement(ruleData,'value')
            value.text="Standard Zone"
            result=SubElement(rule,'result')
            resultName=SubElement(result,'resultName')
            resultName.text=impact_cat
            fieldToValueExpression=SubElement(rule,'fieldToValueExpression')
            operation=SubElement(fieldToValueExpression,'operation')
            operation.text="REGEX"
            seperator=SubElement(fieldToValueExpression,'seperator')
            seperator.text=";"
            fieldName=SubElement(fieldToValueExpression,'fieldName')
            fieldName.text="EventDelayedSessionTelcoGsmSms.RETAIL_ZONE"
            fieldValue=SubElement(fieldToValueExpression,'fieldValue')
            fieldValue.text=zone_imp_cat
            fieldKind=SubElement(fieldToValueExpression,'fieldKind')
            fieldKind.text="EVENT_SPEC_FIELD"
            #############
            fieldToValueExpression=SubElement(rule,'fieldToValueExpression')
            operation=SubElement(fieldToValueExpression,'operation')
            operation.text="REGEX"
            seperator=SubElement(fieldToValueExpression,'seperator')
            seperator.text=";"
            fieldName=SubElement(fieldToValueExpression,'fieldName')
            fieldName.text="EventDelayedSessionTelcoGsmSms.USAGE_CLASS"
            fieldValue=SubElement(fieldToValueExpression,'fieldValue')
            fieldValue.text=usg_clas
            fieldKind=SubElement(fieldToValueExpression,'fieldKind')
            fieldKind.text="EVENT_SPEC_FIELD"
            ##################
            fieldToValueExpression=SubElement(rule,'fieldToValueExpression')
            operation=SubElement(fieldToValueExpression,'operation')
            operation.text="REGEX"
            seperator=SubElement(fieldToValueExpression,'seperator')
            seperator.text=";"
            fieldName=SubElement(fieldToValueExpression,'fieldName')
            fieldName.text="EventDelayedSessionTelcoGsmSms.SERVICE_CLASS"
            fieldValue=SubElement(fieldToValueExpression,'fieldValue')
            fieldValue.text=".*"
            fieldKind=SubElement(fieldToValueExpression,'fieldKind')
            fieldKind.text="EVENT_SPEC_FIELD"
            ###########################
            fieldToValueExpression=SubElement(rule,'fieldToValueExpression')
            operation=SubElement(fieldToValueExpression,'operation')
            operation.text="REGEX"
            seperator=SubElement(fieldToValueExpression,'seperator')
            seperator.text=";"
            fieldName=SubElement(fieldToValueExpression,'fieldName')
            fieldName.text="EventDelayedSessionTelcoGsmSms.SERVICE_CODE"
            fieldValue=SubElement(fieldToValueExpression,'fieldValue')
            fieldValue.text=".*"
            fieldKind=SubElement(fieldToValueExpression,'fieldKind')
            fieldKind.text="EVENT_SPEC_FIELD"

        xml_file=open(pwd + "\data_xml\USC_MMS.xml","w")

        xml_file.write(tostring(top))



def create_mbl_acc_usg_chargmms():
    pwd = os.path.dirname(__file__)
    top=Element('chargeRatePlan')
    top.set('xmlns:pdc','http://xmlns.oracle.com/communications/platform/model/pricing')
    name=SubElement(top,'name')
    name.text="here put the name"
    pricingProfileName=SubElement(top,'pricingProfileName')
    pricingProfileName.text="Convergent Usage"
    priceListName=SubElement(top,'priceListName')
    priceListName.text="Default"
    applicableRums=SubElement(top,'applicableRums')
    applicableRums.text="Occurrence"
    applicableQuantity=SubElement(top,'applicableQuantity')
    applicableQuantity.text="ORIGINAL"
    taxTime=SubElement(top,'taxTime')
    taxTime.text="NONE"
    todMode=SubElement(top,'todMode')
    todMode.text="START_TIME"
    applicableQtyTreatment=SubElement(top,'applicableQtyTreatment')
    applicableQtyTreatment.text="CONTINUOUS"
    permittedName=SubElement(top,'permittedName')
    permittedName.text="TelcoGsm"
    permittedType=SubElement(top,'permittedType')
    permittedType.text="PRODUCT"
    eventName=SubElement(top,'eventName')
    eventName.text="EventDelayedSessionTelcoGsmMms"
    cycleFeeFlag=SubElement(top,'cycleFeeFlag')
    cycleFeeFlag.text="0"
    billOffset=SubElement(top,'billOffset')
    billOffset.text="0"
    subscriberCurrency=SubElement(top,'subscriberCurrency')
    currencyCode=SubElement(subscriberCurrency,'currencyCode')
    currencyCode.text="SEK"
    applicableRum=SubElement(subscriberCurrency,'applicableRum')
    applicableRumName=SubElement(applicableRum,'applicableRumName')
    applicableRumName.text="Occurrence"
    minQuantity=SubElement(applicableRum,'minQuantity')
    minQuantity.text="1.0"
    minQuantityUnit=SubElement(applicableRum,'minQuantityUnit')
    minQuantityUnit.text="NONE"
    incrementQuantity=SubElement(applicableRum,'incrementQuantity')
    incrementQuantity.text="0.0"
    incrementQuantityUnit=SubElement(applicableRum,'incrementQuantityUnit')
    incrementQuantityUnit.text="NONE"
    roundingMode=SubElement(applicableRum,'roundingMode')
    roundingMode.text="NEAREST"
    crpRelDateRange=SubElement(applicableRum,'crpRelDateRange')
    absoluteDateRange=SubElement(crpRelDateRange,'absoluteDateRange')
    startDate=SubElement(absoluteDateRange,'startDate')
    startDate.text="0"
    endDate=SubElement(absoluteDateRange,'endDate')
    endDate.text="inf"
    enhancedZoneModel=SubElement(crpRelDateRange,'enhancedZoneModel')
    zoneModelName=SubElement(enhancedZoneModel,'zoneModelName')
    zoneModelName.text="Standard Zone"
    uscModelName=SubElement(enhancedZoneModel,'uscModelName')
    uscModelName.text="Test SMS USC Selector"
    with open(pwd + '\\temp_text_file\\Mbl_accMobile Access Local MMS.txt') as f:
        for line in f:
            nme=str(line.split(";")[0])

            fp=line.split(";")[1]
            sp=line.split(";")[2]
            gld=line.split(";")[3]
        #    print( str(nme) + "    " + str(sp)  +" " + str(fp) + " " + str(gld) + "\n")
            results=SubElement(enhancedZoneModel,"results")
            name=SubElement(results,"name")
            name.text=str(nme)
            crpCompositePopModel=SubElement(results,"crpCompositePopModel")
            name_crp=SubElement(crpCompositePopModel,"name")
            name_crp.text="Pricing"
            usageChargePopModel=SubElement(crpCompositePopModel,"usageChargePopModel")
            priceTier=SubElement(usageChargePopModel,"priceTier")
            distributionMethod=SubElement(priceTier,"distributionMethod")
            distributionMethod.text="FROM_BAL_IMPACT"
            tierBasis=SubElement(priceTier,"tierBasis")
            rumTierExpression=SubElement(tierBasis,"rumTierExpression")
            enforceCreditLimit=SubElement(priceTier,"enforceCreditLimit")
            enforceCreditLimit.text="false"
            rumName=SubElement(priceTier,"rumName")
            rumName.text="Occurrence"
            priceTierValidityPeriod=SubElement(priceTier,"priceTierValidityPeriod")

            lowerBound=SubElement(priceTierValidityPeriod,"lowerBound")
            lowerBound.text="0"
            validFrom=SubElement(priceTierValidityPeriod,"validFrom")
            validFrom.text="0"
            priceTierRange=SubElement(priceTierValidityPeriod,"priceTierRange")
            upperBound=SubElement(priceTierRange,"upperBound")
            upperBound.text="NO_MAX"

            if fp=="-" or fp == "None":
                print(" ")

            #elif fp !="-" or fp!="None":
            else:
                fixedCharge=SubElement(priceTierRange,"fixedCharge")
                price=SubElement(fixedCharge,"price")
                price.text=str(fp)
                unitOfMeasure=SubElement(fixedCharge,"unitOfMeasure")
                unitOfMeasure.text="NONE"

                balanceElementNumCode=SubElement(fixedCharge,"balanceElementNumCode")
                balanceElementNumCode.text="752"
                discountable=SubElement(fixedCharge,"discountable")
                discountable.text="true"
                priceType=SubElement(fixedCharge,"priceType")
                priceType.text="CONSUMPTION"

                if gld == "None":
                    print("glid is none")
                else:
                    glid=SubElement(fixedCharge,"glid")
                    glid.text=gld

            if sp=="-" or sp == "None":
                print(" ")

            else:
            #if sp!="-" or sp != "None":
                scaledCharge=SubElement(priceTierRange,"scaledCharge")
                price=SubElement(scaledCharge,"price")
                price.text=str(sp)
                unitOfMeasure=SubElement(scaledCharge,"unitOfMeasure")
                unitOfMeasure.text="MINUTES"

                balanceElementNumCode=SubElement(scaledCharge,"balanceElementNumCode")
                balanceElementNumCode.text="752"
                discountable=SubElement(scaledCharge,"discountable")
                discountable.text="true"
                priceType=SubElement(scaledCharge,"priceType")
                priceType.text="CONSUMPTION"

                if gld == "None":
                    print("glid is null")
                else:
                    glid=SubElement(scaledCharge,"glid")
                    glid.text=gld

                incrementStep=SubElement(scaledCharge,"incrementStep")
                incrementStep.text="1"
                incrementRounding=SubElement(scaledCharge,"incrementRounding")
                incrementRounding.text="NONE"



            applicableQuantity=SubElement(priceTier,"applicableQuantity")
            applicableQuantity.text="ORIGINAL"
    xml_file=open(pwd + "\data_xml\mobile_usage_MMS.xml","w")
    xml_file.write(tostring(top))

def create_zone_map():
    pwd = os.path.dirname(__file__)
    top=Element('ns3:ConfigObjects')
    top.set('xmlns:ns2','http://xmlns.oracle.com/communications/platform/model/pricing')
    top.set('xmlns:ns3','http://xmlns.oracle.com/communications/platform/model/Config')
    top.set('xmlns:ns4','http://xmlns.oracle.com/communications/platform/model/Metadata')
    top.set('xmlns:xsi','http://www.w3.org/2001/XMLSchema-instance')
    top.set('xsi:noNamespaceSchemaLocation','ConfigObjects.xsd')
    valuemap=SubElement(top,'valueMap')
    nam=SubElement(valuemap,'name')
    nam.text="IPT Standard Zone"
    description=SubElement(valuemap,'description')
    description.text="Area Codes"
    priceListName=SubElement(valuemap,'priceListName')
    priceListName.text="Default"
    exactMatch=SubElement(valuemap,'exactMatch')
    exactMatch.text="true"
    validityPeriod=SubElement(valuemap,'validityPeriod')
    validFrom=SubElement(validityPeriod,'validFrom')
    validFrom.text="0"
    valueMapKeyo=SubElement(validityPeriod,'valueMapKey')
    keyName=SubElement(valueMapKeyo,'keyName')
    keyName.text="Origin Zone"
    for r in ozdic:
        valueMapKey2=SubElement(valueMapKeyo,'valueMapKey')
        keyName=SubElement(valueMapKey2,'keyName')
        keyName.text=r
        for i in ozdic[str(r)]:
            print("origin" +  i + "\n")
            v=i;
            valueMapValue=SubElement(valueMapKey2,'valueMapValue')
            valueMapValue.text=v

    valueMapKeyd=SubElement(validityPeriod,'valueMapKey')
    keyName=SubElement(valueMapKeyd,'keyName')
    keyName.text="Destination Zone"
    for r in dzdic:
        valueMapKey3=SubElement(valueMapKeyd,'valueMapKey')
        keyName=SubElement(valueMapKey3,'keyName')
        keyName.text=r
        for i in dzdic[str(r)]:
            print("dest" + i + "\n" )
            valueMapValue=SubElement(valueMapKey3,'valueMapValue')
            valueMapValue.text=str(i)

    xml_file=open(pwd + "\data_xml\Zone_map.xml","w")
    xml_file.write(tostring(top))

def create_charge_selector_xml():
    pwd = os.path.dirname(__file__)
    #fil=open('pwd + "\")
    tree = ET.parse(pwd + "\data_xml\CHARGES_PDC.xml")
    root = tree.getroot()
    Moc1=root.findall('chargeRatePlan')
    dict=defaultdict(list)
    for moc in Moc1:
        for node in moc.getiterator():
			    if node.tag=='name':
				            n=node.text

			    if node.tag=='internalId':
				            inid=node.text
				#print("internal id is" + inid)
				            dict[str(n)].append(inid)


    pwd = os.path.dirname(__file__)
    top=Element('pdc:PricingObjectsJXB ')
    top.set('xmlns:pdc','http://xmlns.oracle.com/communications/platform/model/pricing')
    chargeRateplanSelector=SubElement(top,"chargeRateplanSelector")
    chargeRateplanSelector.set('xmlns:pdc','http://xmlns.oracle.com/communications/platform/model/pricing')
    name=SubElement(chargeRateplanSelector,"name")
    name.text="put the name of chaege selector here"
    description=SubElement(chargeRateplanSelector,"description")
    description.text="put discription here"
    internalId=SubElement(chargeRateplanSelector,"internalId")
    internalId.text="put internal id here"
    pricingProfileName=SubElement(chargeRateplanSelector,"pricingProfileName")
    pricingProfileName.text="Convergent Usage"
    priceListName=SubElement(chargeRateplanSelector,"priceListName")
    priceListName.text="Default"
    stereoType=SubElement(chargeRateplanSelector,"stereoType")
    stereoType.text="CHARGE_RATE_PLAN_SELECTOR"
    productSpecName=SubElement(chargeRateplanSelector,"productSpecName")
    productSpecName.text="TelcoGsm"
    eventSpecName=SubElement(chargeRateplanSelector,"eventSpecName")
    eventSpecName.text="EventDelayedSessionTelcoGsmRoaming"
    customerSpecName=SubElement(chargeRateplanSelector,"customerSpecName")
    customerSpecName.text="Account"
    validityPeriod=SubElement(chargeRateplanSelector,"validityPeriod")
    validFrom=SubElement(validityPeriod,"validFrom")
    validFrom.text="0"

    for i in chargelist:
        rule=SubElement(validityPeriod,"rule")
        name=SubElement(rule,"name")
        name.text=i.orizin + "to" + i.destination
        result=SubElement(rule,"result")
        resultName=SubElement(result,"resultName")
        #resultName.text="Roaming " + i.orizin + "to" + i.destination
        resultName.text=i.name + i.glid
        resultData=SubElement(result,"resultData")
        key=SubElement(resultData,"key")
        key.text="RateplanIID"
        value=SubElement(resultData,"value")
        value.text=str(dict[i.name + i.glid])[2:-2]
        resultData=SubElement(result,"resultData")
        key=SubElement(resultData,"key")
        key.text="IMPACT_CATEGORY"
        value=SubElement(resultData,"value")
        value.text=i.orizin + "to" + i.destination
    #    fieldToValueExpression=SubElement(rule,"fieldToValueExpression")
    #    operation=SubElement(fieldToValueExpression,"operation")
    #    operation.text="IN_LIST"
    #    seperator=SubElement(fieldToValueExpression,"seperator")
    #    seperator.text=";"
    #    fieldName=SubElement(fieldToValueExpression,"fieldName")
    #    fieldName.text="VAS01"
    #    fieldKind=SubElement(fieldToValueExpression,"fieldKind")
    #    fieldKind.text="EVENT_SPEC_FIELD"
        valueMapExpression=SubElement(rule,"valueMapExpression")
        operation=SubElement(valueMapExpression,"operation")
        operation.text="MAPS_TO"
        seperator=SubElement(valueMapExpression,"seperator")
        seperator.text=";"
        valueMapName=SubElement(valueMapExpression,"valueMapName")
        valueMapName.text="IPT Standard Zone"
        fieldName=SubElement(valueMapExpression,"fieldName")
        fieldName.text="EventDelayedSessionTelcoGsmRoaming.TELCO_INFO.ORIGIN_NETWORK"
        mappedValue=SubElement(valueMapExpression,"mappedValue")
        orzn=i.orizin.split(".")
        #mappedValue.text="Origin Zone;"+i.orizin
        if len(orzn)==2:
            mappedValue.text="Origin Zone;"+orzn[0]+";"+orzn[1]
        else:
            mappedValue.text="Origin Zone;"+orzn[0]

        fieldKind=SubElement(valueMapExpression,"fieldKind")
        fieldKind.text="EVENT_SPEC_FIELD"
        valueMapExpression=SubElement(rule,"valueMapExpression")
        operation=SubElement(valueMapExpression,"operation")
        operation.text="MAPS_TO"
        seperator=SubElement(valueMapExpression,"seperator")
        seperator.text=";"
        valueMapName=SubElement(valueMapExpression,"valueMapName")
        valueMapName.text="IPT Standard Zone"
        fieldName=SubElement(valueMapExpression,"fieldName")
        fieldName.text="EventDelayedSessionTelcoGsmRoaming.TELCO_INFO.DESTINATION_NETWORK"
        mappedValue=SubElement(valueMapExpression,"mappedValue")
        dst=i.destination.split(".")
        if len(dst)==2:
            mappedValue.text="Destination Zone;" + dst[0]+";"+dst[1]
        else:
            mappedValue.text="Destination Zone;" + dst[0]+";"

        fieldKind=SubElement(valueMapExpression,"fieldKind")
        fieldKind.text="EVENT_SPEC_FIELD"

    xml_file=open(pwd + "\data_xml\charge_selector.xml","w")
    xml_file.write(tostring(top))






def create_charges():
    pwd = os.path.dirname(__file__)
    top=Element('pdc:PricingObjectsJXB ')
    top.set('xmlns:pdc','http://xmlns.oracle.com/communications/platform/model/pricing')
    for i in chrgdict:
        n=str(i)
    #for i in chargelist:
        end=len(n)-2
        prc=n[7:end]
        l=set()
        for r in chrgdict[str(i)]:
            l.add(r)
        print("printing values after appending to list" + n)
        for x in l:
            print x
        for x in l:
            chrgrateplan=SubElement(top,'chargeRatePlan')
            chrgrateplan.set('xmlns:pdc','http://xmlns.oracle.com/communications/platform/model/pricing')
            name=SubElement(chrgrateplan,'name')
            #name.text=str(i.name)
            name.text=str(n) + x
            pricingProfileName=SubElement(chrgrateplan,'pricingProfileName')
            pricingProfileName.text="Convergent Usage"
            priceListName=SubElement(chrgrateplan,'priceListName')
            priceListName.text="Default"
            applicableRums=SubElement(chrgrateplan,'applicableRums')
            applicableRums.text="Duration"
            applicableQuantity=SubElement(chrgrateplan,'applicableQuantity')
            applicableQuantity.text="ORIGINAL"
            taxTime=SubElement(chrgrateplan,'taxTime')
            taxTime.text="NONE"
            todMode=SubElement(chrgrateplan,'todMode')
            todMode.text="START_TIME"
            applicableQtyTreatment=SubElement(chrgrateplan,'applicableQtyTreatment')
            applicableQtyTreatment.text="CONTINUOUS"
            permittedName=SubElement(chrgrateplan,'permittedName')
            permittedName.text="TelcoGsm"
            permittedType=SubElement(chrgrateplan,'permittedType')
            permittedType.text="PRODUCT"
            eventName=SubElement(chrgrateplan,'eventName')
            eventName.text="EventDelayedSessionTelcoGsmRoaming"
            cycleFeeFlag=SubElement(chrgrateplan,'cycleFeeFlag')
            cycleFeeFlag.text="0";
            billOffset=SubElement(chrgrateplan,'billOffset')
            billOffset.text="0"
            subscriberCurrency=SubElement(chrgrateplan,'subscriberCurrency')
            currencyCode=SubElement(subscriberCurrency,'currencyCode')
            currencyCode.text="SEK"
            applicableRum=SubElement(subscriberCurrency,'applicableRum')
            applicableRumName=SubElement(applicableRum,'applicableRumName')
            applicableRumName.text="Duration"
            minQuantity=SubElement(applicableRum,'minQuantity')
            minQuantity.text="1.0"
            minQuantityUnit=SubElement(applicableRum,'minQuantityUnit')
            minQuantityUnit.text="NONE"
            incrementQuantity=SubElement(applicableRum,'incrementQuantity')
            incrementQuantity.text="0.0"
            incrementQuantityUnit=SubElement(applicableRum,'incrementQuantityUnit')
            incrementQuantityUnit.text="NONE"
            roundingMode=SubElement(applicableRum,'roundingMode')
            roundingMode.text="NEAREST"
            crpRelDateRange=SubElement(applicableRum,'crpRelDateRange')
            absoluteDateRange=SubElement(crpRelDateRange,'absoluteDateRange')
            startDate=SubElement(absoluteDateRange,'startDate')
            startDate.text="0"
            endDate=SubElement(absoluteDateRange,'endDate')
            endDate.text="inf"
            crpCompositePopModel=SubElement(crpRelDateRange,'crpCompositePopModel')
            name=SubElement(crpCompositePopModel,'name')
            name.text="Pricing"
            usageChargePopModel=SubElement(crpCompositePopModel,'usageChargePopModel')
            priceTier=SubElement(usageChargePopModel,'priceTier')
            distributionMethod=SubElement(priceTier,'distributionMethod')
            distributionMethod.text="FROM_BAL_IMPACT"
            tierBasis=SubElement(priceTier,'tierBasis')
            rumTierExpression=SubElement(tierBasis,'rumTierExpression')
            enforceCreditLimit=SubElement(priceTier,'enforceCreditLimit')
            enforceCreditLimit.text="false"
            rumName=SubElement(priceTier,'rumName')
            rumName.text="Duration"
            priceTierValidityPeriod=SubElement(priceTier,'priceTierValidityPeriod')
            lowerBound=SubElement(priceTierValidityPeriod,'lowerBound')
            lowerBound.text="0"
            validFrom=SubElement(priceTierValidityPeriod,'validFrom')
            validFrom.text="0"
            priceTierRange=SubElement(priceTierValidityPeriod,'priceTierRange')
            upperBound=SubElement(priceTierRange,'upperBound')
            upperBound.text="NO_MAX"
            scaledCharge=SubElement(priceTierRange,'scaledCharge')
            price=SubElement(scaledCharge,'price')
            price.text=prc
            unitOfMeasure=SubElement(scaledCharge,'unitOfMeasure')
            unitOfMeasure.text="MINUTES"
            balanceElementNumCode=SubElement(scaledCharge,'balanceElementNumCode')
            balanceElementNumCode.text="752"
            discountable=SubElement(scaledCharge,'discountable')
            discountable.text="true"
            priceType=SubElement(scaledCharge,'priceType')
            priceType.text="CONSUMPTION"
            glid=SubElement(scaledCharge,'glid')
            #glid.text=str(i.glid)
            glid.text=x
            incrementStep=SubElement(scaledCharge,'incrementStep')
            incrementStep.text="1"
            incrementRounding=SubElement(scaledCharge,'incrementRounding')
            incrementRounding.text="NONE"
            applicableQuantity=SubElement(priceTier,'applicableQuantity')
            applicableQuantity.text="ORIGINAL"

    xml_file=open(pwd + "\data_xml\charges.xml","w")
    xml_file.write(tostring(top))


def cleanup():
    folder = 'C:\\Users\\chankaya.singh\\Desktop\\repository\\pdc_config_latest\\pdc_configurations\\media\\xl'
    for the_file in os.listdir(folder):
        file_path = os.path.join(folder, the_file)
        try:
            if os.path.isfile(file_path):
                os.unlink(file_path)
        #elif os.path.isdir(file_path): shutil.rmtree(file_path)
        except Exception as e:
            print(e)
