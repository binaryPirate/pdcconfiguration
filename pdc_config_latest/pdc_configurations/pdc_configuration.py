import sys
import openpyxl
import os
from openpyxl import load_workbook
from xml.etree.ElementTree import Element, tostring

from xml.etree.ElementTree import Element, SubElement, Comment, tostring
import xml.dom.minidom


reload(sys)
sys.setdefaultencoding('utf8')
#open("C:\Users\chankaya.singh\Desktop\python_tool\backup_pdc\latest\pdc_config_latest\temp_text_file\zone_model.txt", 'w').close()
#open("xml_zone_model.xml","w").close()
#open("C:\Users\chankaya.singh\Desktop\python_tool\backup_pdc\latest\pdc_config_latest\temp_text_file\balance_element.txt","w").close()
#open("C:\Users\chankaya.singh\Desktop\python_tool\backup_pdc\latest\pdc_config_latest\temp_text_file\temp_text_file\impact_category.txt","w").close()
#open("xml/xml_balance_element.xml","w").close()
#open("xml/xml_impact_cat.xml","w").close()

name_of_configuration=sys.argv[2]

def get_data():

    source_file=str(sys.argv[1])
    wb=load_workbook(filename=source_file)
    sheet_ranges=wb[str(sys.argv[2])]
    row=sheet_ranges.max_row
    if name_of_configuration == 'Zone Model':
        file=open("temp_text_file/Zone_Model.txt","a")
        for r in range(3,row):
            des=str(sheet_ranges['G' + str(r)].value)
            nam=str(sheet_ranges['H' + str(r)].value)
            file.write( des + ";" + nam + "\n")
        file.close()
    elif name_of_configuration == 'Balance Elements':
        open("temp_text_file/balance_element.txt","w").close()
        file=open("temp_text_file/balance_element.txt","w")

        row=row+1
        for r in range(3,row):

            name=sheet_ranges['B' + str(r)].value



            code=str(sheet_ranges['C' + str(r)].value)

            numericcode=str(sheet_ranges['D' + str(r)].value)
            symbol=str(sheet_ranges['E' + str(r)].value)

            pr1=str(sheet_ranges['O' + str(r)].value)

            tolmin1=str(sheet_ranges['P' + str(r)].value)

            tolmax1=str(sheet_ranges['Q' + str(r)].value)

            tolper1=str(sheet_ranges['R' + str(r)].value)

            round1=str(sheet_ranges['N' + str(r)].value)

            pr2=str(sheet_ranges['x' + str(r)].value)

            tolmin2=str(sheet_ranges['Y' + str(r)].value)

            tolmax2=str(sheet_ranges['Z' + str(r)].value)

            tolper2=str(sheet_ranges['AA' + str(r)].value)

            round2=str(sheet_ranges['W' + str(r)].value)

            pr3=str(sheet_ranges['AG' + str(r)].value)

            tolmin3=str(sheet_ranges['AH' + str(r)].value)

            tolmax3=str(sheet_ranges['AI' + str(r)].value)

            tolper3=str(sheet_ranges['AJ' + str(r)].value)

            round3=str(sheet_ranges['AF' + str(r)].value)

            pr4=str(sheet_ranges['AP' + str(r)].value)

            tolmin4=str(sheet_ranges['AQ' + str(r)].value)

            tolmax4=str(sheet_ranges['AR' + str(r)].value)

            tolper4=str(sheet_ranges['AS' + str(r)].value)

            round4=str(sheet_ranges['AO' + str(r)].value)
            #file.write(name)
            file.write(name + ";" + "Default" + ";" + code + ";" + numericcode + ";" + symbol + ";" + pr1 + ";" + tolmin1 + ";" + tolmax1 + ";" + tolper1 + ";" + round1 + ";" + pr2 + ";" + tolmin2 + ";" + tolmax2 + ";" + tolper2 + ";" + round2 + ";" + pr3 + ";" + tolmin3 + ";" + tolmax3 + ";" + tolper3 + ";" + round3 + ";" + pr4 + ";" + tolmin4 + ";" + tolmax4 + ";" + tolper4 + ";" + round4 + ";"  + "\n")

	file.close()
    elif name_of_configuration == 'Impact Category':
         file=open("temp_text_file/impact_category.txt","a")
         for r in range(3,row):
             name=str(sheet_ranges['B' + str(r)].value)
             descr=str(sheet_ranges['C' + str(r)].value)
             s=len(descr)
             if descr[0] == "=":
                ind=descr[1:5]
                if descr[s-4:s-1] == "IDD" :
                    descr=str(sheet_ranges[str(ind)].value)+" " + str(descr[s-4:s-1])
                  #  print(descr + " ")
                else:
                    descr=str(sheet_ranges[str(ind)].value)
                   # print(descr + "\n")



             result=name
             file.write(name + ";" + descr + ";" + result + "\n")
         file.close()


    else:
         print(" pass the right configuration")



def create_xml_zone_model():


    top=Element('cim:ConfigObjects')
    top.set('xmlns:cim','http://xmlns.oracle.com/communications/platform/model/Config')

    standardZoneModel=SubElement(top,'standardZoneModel')
    standardZoneModel.set('xmlns:mtd','http://xmlns.oracle.com/communications/platform/model/Metadata')
    standardZoneModel.set('xmlns:cim','http://xmlns.oracle.com/communications/platform/model/Config')
    standardZoneModel.set('xmlns:pdc','http://xmlns.oracle.com/communications/platform/model/pricing')
    name=SubElement(standardZoneModel,'name')
    name.text='Standard Zone'
    description=SubElement(standardZoneModel,'description')
    description.text='Standard Zone'
    priceListName=SubElement(standardZoneModel,'priceListName')
    priceListName.text='Default'
    obsolete=SubElement(standardZoneModel,'obsolete')
    obsolete.text='false'


    with open("temp_text_file/zone_model.txt") as f:
		for line in f:
			desc=line.split(';')[0]
			name=line.split(';')[1]
			name=name.strip()
			desc=desc.strip()
		#	print(name + " " + desc)
			zoneItem=SubElement(standardZoneModel,'zoneItem')
			productName=SubElement(zoneItem,'productName')
			productName.text='*'
			originPrefix=SubElement(zoneItem,'originPrefix')
			originPrefix.text='0'
			destinationPrefix=SubElement(zoneItem,'destinationPrefix')
			destinationPrefix.text=desc
			validFrom=SubElement(zoneItem,'validFrom')
			validFrom.text='0'
			validTo=SubElement(zoneItem,'validTo')
			validTo.text='inf'
			zoneResult=SubElement(zoneItem,'zoneResult')
			zoneName=SubElement(zoneResult,'zoneName')
			zoneName.text=name


    xml_file=open("data_xml/Zone_Model.xml","w")
    #xml_t=xml.dom.minidom.parse(top)
    #pretty_xml=xml_t.toprettyxml()
    #print(pretty_xml)
    xml_file.write(tostring(top))

def create_balance_element():
    top=Element('cim:ConfigObjects')
    top.set('xmlns:cim','http://xmlns.oracle.com/communications/platform/model/Config')
  #  top.set('xmlns:pdc','http://xmlns.oracle.com/communications/platform/model/pricing')
    with open("temp_text_file/balance_element.txt") as f:
		for line in f:
                        nam=line.split(';')[0]
                       # nam=nam.strip()
                        cod=line.split(';')[2]
                        numericcode=line.split(';')[3]
                        symbl=line.split(';')[4]
                        pr1=line.split(';')[5]
                        tolmin1=line.split(';')[6]

                        tolmax1=line.split(';')[7]

                        tolper1=line.split(';')[8]
                        if tolmin1 == "0":
                            tolmin1="0.0"

                        if tolper1 == "0":
                            tolper1="0.0"

                        if tolmax1 == "0":
                            tolmax1="0.0"
                        round1=line.split(';')[9]
                        round1=round1.upper()

                        pr2=line.split(';')[10]
                        tolmin2=line.split(';')[11]
                        tolmax2=line.split(';')[12]
                        tolper2=line.split(';')[13]
                        if tolmin2 == "0":
                            tolmin2="0.0"

                        if tolper2 == "0":
                            tolper2="0.0"

                        if tolmax2 == "0":
                            tolmax2="0.0"
                        round2=line.split(';')[14]
                        round2=round2.upper()

                        pr3=line.split(';')[15]
                        tolmin3=line.split(';')[16]
                        tolmax3=line.split(';')[17]
                        tolper3=line.split(';')[18]

                        if tolmin3 == "0":
                            tolmin3="0.0"

                        if tolper3 == "0":
                            tolper3="0.0"

                        if tolmax3 == "0":
                            tolmax3="0.0"
                        round3=line.split(';')[19]
                        round3=round3.upper()

                        pr4=line.split(';')[20]
                        tolmin4=line.split(';')[21]
                        tolmax4=line.split(';')[22]
                        tolper4=line.split(';')[23]
                        if tolmin4 == "0":
                            tolmin4="0.0"

                        if tolper4 == "0":
                            tolper4="0.0"

                        if tolmax4 == "0":
                            tolmax4="0.0"
                        round4=line.split(';')[24]
                        round4=round4.upper()
                        #print(round1.upper() + " " + round2 + " " + round3 + " " + round4 + "\n")

			balanceElements=SubElement(top,'balanceElements')
			balanceElements.set('xmlns:mtd','http://xmlns.oracle.com/communications/platform/model/Metadata')
			balanceElements.set('xmlns:cim','http://xmlns.oracle.com/communications/platform/model/Config')
			balanceElements.set('xmlns:pdc','http://xmlns.oracle.com/communications/platform/model/pricing')
			name=SubElement(balanceElements,'name')
			name.text=nam;
			description=SubElement(balanceElements,'description')
			description.text=nam;
			priceListName=SubElement(balanceElements,'priceListName')
			priceListName.text="default"
			obsolete=SubElement(balanceElements,'obsolete')
			obsolete.text="false"
			code=SubElement(balanceElements,'code')
			code.text=cod
			numericCode=SubElement(balanceElements,'numericCode')
			numericCode.text=numericcode
			symbol=SubElement(balanceElements,'symbol')
			symbol.text=symbl
			transientElement=SubElement(balanceElements,'transientElement')
			transientElement.text="false"
			foldable=SubElement(balanceElements,'foldable')
			foldable.text="true"
			counter=SubElement(balanceElements,'counter')
			counter.text="false"
			roundingRules=SubElement(balanceElements,'roundingRules')
			precision=SubElement(roundingRules,'precision')
			precision.text=pr1
			toleranceMin=SubElement(roundingRules,'toleranceMin')
			toleranceMin.text=tolmin1
			toleranceMax=SubElement(roundingRules,'toleranceMax')
			toleranceMax.text=tolmax1
			tolerancePercentage=SubElement(roundingRules,'tolerancePercentage')
			tolerancePercentage.text=tolper1
			roundingMode=SubElement(roundingRules,'roundingMode')
			roundingMode.text=round1
			processingStage=SubElement(roundingRules,'processingStage')
                        processingStage.text='CHARGING'

                        roundingRules=SubElement(balanceElements,'roundingRules')
			precision=SubElement(roundingRules,'precision')
			precision.text=pr2
			toleranceMin=SubElement(roundingRules,'toleranceMin')
			toleranceMin.text=tolmin2
			toleranceMax=SubElement(roundingRules,'toleranceMax')
			toleranceMax.text=tolmax2
			tolerancePercentage=SubElement(roundingRules,'tolerancePercentage')
			tolerancePercentage.text=tolper2
			roundingMode=SubElement(roundingRules,'roundingMode')
			roundingMode.text=round2
			processingStage=SubElement(roundingRules,'processingStage')
                        processingStage.text='ALTERATION'

			roundingRules=SubElement(balanceElements,'roundingRules')
			precision=SubElement(roundingRules,'precision')
			precision.text=pr3
			toleranceMin=SubElement(roundingRules,'toleranceMin')
			toleranceMin.text=tolmin3
			toleranceMax=SubElement(roundingRules,'toleranceMax')
			toleranceMax.text=tolmax3
			tolerancePercentage=SubElement(roundingRules,'tolerancePercentage')
			tolerancePercentage.text=tolper3
			roundingMode=SubElement(roundingRules,'roundingMode')
			roundingMode.text=round3
			processingStage=SubElement(roundingRules,'processingStage')
                        processingStage.text='TAXATION'

			roundingRules=SubElement(balanceElements,'roundingRules')
			precision=SubElement(roundingRules,'precision')
			precision.text=pr4
			toleranceMin=SubElement(roundingRules,'toleranceMin')
			toleranceMin.text=tolmin4
			toleranceMax=SubElement(roundingRules,'toleranceMax')
			toleranceMax.text=tolmax4
			tolerancePercentage=SubElement(roundingRules,'tolerancePercentage')
			tolerancePercentage.text=tolper4
			roundingMode=SubElement(roundingRules,'roundingMode')
			roundingMode.text=round4
			processingStage=SubElement(roundingRules,'processingStage')
                        processingStage.text='AR'

			consumptionRule=SubElement(balanceElements,'consumptionRule')
			consumptionRule.text="NONE"


	#xml=xml.dom.minidom.parse(top)
	#pretty_xml_as_string=xml.toprettyxnl()
	#print(pretty_xml_as_string)

    xml_file=open("data_xml/Balance_Elements.xml","w")
    xml_file.write(tostring(top))

# function to create xml for impact category
def create_impact_category():
    top=Element('cim:ConfigObjects')
    top.set('xmlns:cim','http://xmlns.oracle.com/communications/platform/model/Config')
    with open("temp_text_file/impact_category.txt") as f:
        for line in f:

            nam=line.split(";")[0]
            descr=line.split(";")[1]
            reslt=line.split(";")[2]
            zoneResultConfiguration=SubElement(top,'zoneResultConfiguration')
            zoneResultConfiguration.set('xmlns:mtd','http://xmlns.oracle.com/communications/platform/model/Metadata')
            zoneResultConfiguration.set('xmlns:cim','http://xmlns.oracle.com/communications/platform/model/Config')
            zoneResultConfiguration.set('xmlns:pdc','http://xmlns.oracle.com/communications/platform/model/pricing')
            name=SubElement(zoneResultConfiguration,'name')
            name.text=nam
            description=SubElement(zoneResultConfiguration,'description')
            description.text=descr
            priceListName=SubElement(zoneResultConfiguration,'priceListName')
            priceListName.text="Default"
            obsolete=SubElement(zoneResultConfiguration,'obsolete')
            obsolete.text="false"
            result=SubElement(zoneResultConfiguration,'result')
            result.text=reslt
            resultType=SubElement(zoneResultConfiguration,'resultType')
            resultType.text="BASE_RESULT"
        xml_file=open("data_xml/Impact_Category.xml","w")

        xml_file.write(tostring(top))






if name_of_configuration == 'Zone Model':
    get_data()
    create_xml_zone_model();
   # cmd="python C:\Users\chankaya.singh\Desktop\python_tool\django\pdc_config\pdc_configurations\pdc_configuration.py" + " " + "\"" + newest + "\" " + " " + "\"" + str(config_name) + "\"" ""
  #  pscpcmd='C:\\"Program Files"\\putty\\pscp.exe -pw "Gh4df!18st#" C:\Users\chankaya.singh\Desktop\python_tool\django\pdc_config\data_xml\\zone_model.xml  brm@tr005stbrm.ddc.teliasonera.net:/opt/brm/brm/portal/7.5/chan C:\Users\chankaya.singh\Desktop\python_tool\django\pdc_config\data_xml\\zone_model.xml  brm@tr005stbrm.ddc.teliasonera.net:/opt/brm/brm/portal/7.5/chan'

   # os.system(pscpcmd)
    #cmd='C:\\"Program Files"\\putty\\plink.exe brm@tr005stbrm.ddc.teliasonera.net -pw Gh4df!18st# -m C:\Users\chankaya.singh\Desktop\\temp\cd.txt'
    #os.system(cmd)
elif name_of_configuration == 'Balance Elements':
    get_data()
    create_balance_element()
    #pscpcmd='C:\\"Program Files"\\putty\\pscp.exe -pw "Gh4df!18st#" C:\Users\chankaya.singh\Desktop\python_tool\django\pdc_config\data_xml\\balance_element.xml  brm@tr005stbrm.ddc.teliasonera.net:/opt/brm/brm/portal/7.5/chan'

    #os.system(pscpcmd)
elif name_of_configuration == 'Impact Category':
    get_data()
    create_impact_category();
    #pscpcmd='C:\\"Program Files"\\putty\\pscp.exe -pw "Gh4df!18st#" C:\Users\chankaya.singh\Desktop\python_tool\django\pdc_config\data_xml\\impact_cat.xml  brm@tr005stbrm.ddc.teliasonera.net:/opt/brm/brm/portal/7.5/chan'

    #os.system(pscpcmd)
else:
    print("please pass proper configuration name")
